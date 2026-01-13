import asyncio

from openpyxl import load_workbook, Workbook
import os

import sqlite_comands as sql
from logger import logger
from settings import WORK_CNT_DB, POL_MEL_DB


async def line_breaks(dir_path_):
    print(dir_path_)
    wb = load_workbook(dir_path_)
    ws = wb.active

    col_num = 1
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=10):
        for cell in row:
            col_num += 1
            value = cell.value
            if not value or 'писание' in value:
                break
        if col_num == 9:
            return 'Не найдена колонка "Описание"'

    max_row = ws.max_row

    for row in range(2, max_row+1):
        old_value = ws.cell(row=row, column=col_num-1).value
        if not old_value:
            continue
        new_value = old_value.replace('•', ' <br> •').replace('<br>  <br> •', ' <br> •')
        letter_idx = 0
        for letter in new_value:
            if letter == ' ' or letter == '<' or letter == 'b' or letter == 'r' or letter == '>':
                letter_idx += 1
            else:
                break
        new_value = new_value[letter_idx:]
        ws.cell(row=row, column=col_num-1).value = new_value

    wb.save(dir_path_)


async def cards_count(dir_path_):
    """Возвращает количество карточек минимальной и максимальной стоимости"""
    wb = load_workbook(dir_path_)
    ws = wb.active

    column_name, column_desc, column_pic = False, False, False
    col_num = 1
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=10):
        for cell in row:
            col_num += 1
            value = cell.value

            if not value:
                break

            if 'Наименование' in value:
                column_name = col_num
            elif 'Описание' in value:
                column_desc = col_num
            elif 'фото' in value:
                column_pic = col_num
    if not column_name or not column_desc or not column_pic:
        print(f"column_name - {column_name}, column_desc - {column_desc}, column_pic - {column_pic},")
        return 'Не найдена одна из колонок, "Описание", "Наименование", "Ссылки на фото"'

    max_row = ws.max_row
    min_cost, max_cost, zero_cost = 0, 0, 0
    for row in range(1, max_row+1):
        description = ws.cell(row=row+1, column=column_desc-1).value
        pic_cnt = ws.cell(row=row+1, column=column_pic-1).value

        if not description and not pic_cnt:
            if not ws.cell(row=row+1, column=1).value:
                break
            else:
                zero_cost += 1
                continue
        else:
            if pic_cnt and description:
                cnt_str = description.count('•')
                cnt_photo = len(pic_cnt.split(';'))
                if cnt_str >= 3 and cnt_photo >= 2:
                    max_cost += 1
                else:
                    min_cost += 1
            else:
                print(row)

    return min_cost, max_cost


async def work_cnt(cnt_table_name):
    wb = Workbook()
    ws = wb.active

    column = 2
    for name in ['Исполнитель', 'Неполных', 'Полных', 'Стоимость неполных', 'Стоимость полных', 'Итого должны']:
        ws.cell(row=2, column=column).value = name
        column += 1

    column = 2
    for name in ['Название', 'Неполных', 'Полных', 'Исполнитель']:
        ws.cell(row=13, column=column).value = name
        column += 1

    DB = sql.DBMagic(POL_MEL_DB)
    user_list = await DB.check_exist("Users", id_="All")
    table_data_list = await DB.get_unverified_table()
    logger.debug(f"table_data_list: {table_data_list} \nuser_list: {user_list}")

    row_table_cnt = 14
    row_work_cnt = 3
    for user in user_list:
        logger.debug(f"user_list: {user_list}")
        tg_id, user_name = user[0], user[1]
        logger.debug(f"tg_id: {tg_id}\n, user_name: {user_name}")

        incomplete_cards_cnt = 0
        complete_cards_cnt = 0
        for table in table_data_list:
            table_name, incomplete_cards, complete_cards = table[1], table[3], table[4]
            logger.debug(f"table_name: {table_name}\n, (in)complete_cards: {incomplete_cards}/{complete_cards}")
            logger.debug(f"tg_id: {table[5]}")
            if table[5] == tg_id:
                try:
                    incomplete_cards_cnt += int(incomplete_cards)
                except Exception as e:
                    logger.error(f"Ошибка при обработке таблицы {table_name}\ntable_data_list - {table_data_list} \n")
                    raise e
                complete_cards_cnt += int(complete_cards)

                ws.cell(row=row_table_cnt, column=2).value = table_name
                ws.cell(row=row_table_cnt, column=3).value = incomplete_cards
                ws.cell(row=row_table_cnt, column=4).value = complete_cards
                ws.cell(row=row_table_cnt, column=5).value = user_name

                row_table_cnt += 1
            else:
                continue

        ws.cell(row=row_work_cnt, column=2).value = user_name
        ws.cell(row=row_work_cnt, column=3).value = incomplete_cards_cnt
        ws.cell(row=row_work_cnt, column=4).value = complete_cards_cnt
        ws.cell(row=row_work_cnt, column=5).value = incomplete_cards_cnt * 6
        ws.cell(row=row_work_cnt, column=6).value = complete_cards_cnt * 9
        ws.cell(row=row_work_cnt, column=7).value = (incomplete_cards_cnt * 6) + (complete_cards_cnt * 9)
        row_work_cnt += 1

    file_path = os.path.join('tables', cnt_table_name)
    wb.save(file_path)
    col_num = 1
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=10):
        for cell in row:
            col_num += 1
            value = cell.value
            if not value or 'писание' in value:
                break
        if col_num == 9:
            return 'Не найдена колонка "Описание"'


async def name_update(table_name):
    wb = load_workbook(table_name)
    ws = wb.active

    column_name, column_code = 0, 0
    col_num = 0
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=10):
        for cell in row:
            col_num += 1
            value = cell.value
            if not value:
                break

            if 'Наименование' in value:
                column_name = col_num
            elif 'Код' in value:
                column_code = col_num
    if not column_name or not column_code:
        logger.error(f'Не найдена одна из колонок. column_name - {column_name}, column_code - {column_code}')
        return None

    cnt_change, zero_cnt = 0, 0
    for row in range(2, ws.max_row+1):
        code, name = ws.cell(row=row, column=column_code).value, ws.cell(row=row, column=column_name).value
        if code == " " or not code:
            ws.delete_rows(idx=row, amount=1)
            continue

        prod_db = sql.DBMagic(POL_MEL_DB)
        product_data = await prod_db.check_exist("products_product", "sbis_id", code)
        if product_data:
            name_from_db = product_data[0][2]

            if name_from_db != name:
                ws.cell(row=row, column=column_name).value = name_from_db
                cnt_change += 1
                logger.debug(f'В таблице обновлено наименование. Было: {name} Стало: {name_from_db}')
        else:
            ws.delete_rows(idx=row, amount=1)
            logger.warning(f'Товара {name} нет в БД. Он удалем из таблицы.')
            zero_cnt += 1
            continue
    logger.warning(f"При обработке таблицы {table_name}, было изменено {cnt_change}, не найдено: {zero_cnt}")
    wb.save(table_name)
    return f"При обработке таблицы {table_name}, было изменено {cnt_change}, не найдено: {zero_cnt}"


async def price_update(table_name):
    wb = load_workbook(table_name)
    ws = wb.active

    "Определяем поля"
    price_cols = []
    price_fore_site, code_col = None, None
    for num, col in enumerate(ws.iter_cols(min_row=1, max_row=1, values_only=True)):
        if 'Прайс' in col[0]:
            price_cols.append(num)
        elif '!!!ДЛЯ САЙТА!!!' in col[0]:
            price_fore_site = num + 1
        elif 'Код' in col[0]:
            code_col = num
    if not price_fore_site or not price_cols:
        return "Таблица не подходит для обновления цен. Отсутствует одно из полей."

    "Итерируемся по каждой строке таблицы"
    for num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
        if row[code_col] == ' ' or row[code_col] == '' or not row[code_col]:
            ws.delete_rows(num+2)
            continue
        product_price_list = [row[price] for price in price_cols if isinstance(row[price], (int, float))]
        #Узнали минимальную стоимость и обновили колонку с ценой
        if not product_price_list:
            ws.delete_rows(num+2)
            continue
        max_cost = max(product_price_list)
        ws.cell(row=num+2, column=price_fore_site).value = max_cost
    wb.save(table_name)
    return True


if __name__ == "__main__":
    asyncio.run(line_breaks("Светильники (1).xlsx"))
